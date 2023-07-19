from openpyxl import Workbook  
import time  
  
wb = Workbook()  
sheet = wb.active  
  
sheet['A1'] = 87  
sheet['A2'] = "Devansh"  
sheet['A3'] = 41.80  
sheet['A4'] = 10  
sheet['A9'] = 29 
now = time.strftime("%x")  
sheet['A5'] = now  
  
wb.save("sample_file.xlsx") 


# -----------------------------------------------------------------------
# Excel Operation

# from openpyxl import workbook, load_workbook

# # wb = workbook()
# wb = load_workbook("sample_file.xlsx")
# sheet = wb.active
# wb.save("sample_file.xlsx") 
# print(wb.get_sheet_name())
# sheet = wb["emp_data"]





